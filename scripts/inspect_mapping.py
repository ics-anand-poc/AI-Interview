import openpyxl

wb = openpyxl.load_workbook('Product Mapping details 18th July 26.xlsx', data_only=True)
print("Sheets in Product Mapping details:", wb.sheetnames)
for name in wb.sheetnames:
    sheet = wb[name]
    rows = list(sheet.iter_rows(values_only=True))
    print(f"Sheet {name}: header = {rows[0] if rows else None}, row count = {len(rows)}")
    for r in rows[1:10]:
        print("  ", r)
