import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
import re

def normalize(s):
    if s is None:
        return ''
    s = str(s).strip()
    s = re.sub(r'[^a-zA-Z0-9]+', ' ', s)
    s = re.sub(r'\s+', ' ', s)
    return s.lower().strip()

wb2 = openpyxl.load_workbook('Question Bank-20th July \'26.xlsx', data_only=True)
qb_products = {}

for sheet_name in wb2.sheetnames:
    sheet = wb2[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        continue
    header = [str(c).strip() if c else '' for c in rows[0]]
    
    prod_col = None
    q_col = None
    for i, h in enumerate(header):
        h_norm = normalize(h)
        if 'product' in h_norm and prod_col is None:
            prod_col = i
        if 'question' in h_norm and q_col is None:
            q_col = i
            
    prods_in_sheet = set()
    for row in rows[1:]:
        prod_val = None
        if prod_col is not None and prod_col < len(row):
            prod_val = row[prod_col]
        if not prod_val:
            prod_val = sheet_name
        if prod_val:
            prods_in_sheet.add(str(prod_val).strip())
            
    print(f'Sheet [{sheet_name}]: Products found = {sorted(list(prods_in_sheet))[:5]} (Total: {len(prods_in_sheet)})')

wb1 = openpyxl.load_workbook('Resource details less tahn 3.5 rating.xlsx', data_only=True)
sheet1 = wb1.active
emp_rows = list(sheet1.iter_rows(values_only=True))

emp_prods = set()
for r in emp_rows[1:]:
    p = r[4]
    if p:
        emp_prods.add(str(p).strip())

print("\n--- ALL EMP PRODUCTS vs SHET NAMES & PROD COLS ---")
for ep in sorted(list(emp_prods)):
    print(f"Emp product: {ep} | Normalized: {normalize(ep)}")
