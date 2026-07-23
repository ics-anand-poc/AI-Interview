import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
import random
import re
import os

def clean_str(val):
    if val is None:
        return ""
    return str(val).strip()

def normalize_token(s):
    if not s:
        return ""
    s = s.upper().strip()
    s = re.sub(r'[^A-Z0-9]+', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s

def get_product_key(raw_name, sheet_name=""):
    t_raw = normalize_token(raw_name)
    t_sheet = normalize_token(sheet_name)
    
    for t in [t_raw, t_sheet]:
        if t in ['HLR HSS', 'HSS HLR', 'HLR HSS CNF DEPLOYMENT', 'HLR HSS CNF', 'HSS', 'HLR']:
            return 'HLR-HSS'
        if t in ['AUSF UDM', 'UDM']:
            return 'UDM'
        if t in ['SDL DEPLOYMENT AND UPGRADE', 'SDL']:
            return 'SDL'
        if t in ['AAA', 'CSD', 'NPC', 'NCC', 'CMG', 'CMM', 'NRD', 'CFX', 'MRF', 'NN', 'CBIS', 'NCOM', 'NCP', 'CBAM', 'NCD']:
            return t
    return t_raw if t_raw else t_sheet

# 1. Load Question Bank
print("Loading Question Bank...")
wb_qb = openpyxl.load_workbook("Question Bank-20th July '26.xlsx", data_only=True)
qb_by_product = {}

for sheetname in wb_qb.sheetnames:
    sheet = wb_qb[sheetname]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        continue
    header = [clean_str(c) for c in rows[0]]
    prod_col = None
    q_col = None
    for i, h in enumerate(header):
        h_lower = h.lower()
        if 'product' in h_lower and prod_col is None:
            prod_col = i
        if 'question' in h_lower and q_col is None:
            q_col = i
            
    for r in rows[1:]:
        q_val = clean_str(r[q_col]) if q_col is not None and q_col < len(r) else ""
        if not q_val:
            continue
        prod_val = clean_str(r[prod_col]) if prod_col is not None and prod_col < len(r) else ""
        
        key = get_product_key(prod_val, sheetname)
        if key not in qb_by_product:
            qb_by_product[key] = []
        qb_by_product[key].append(q_val)

# Deduplicate questions per product key while preserving order
for key in qb_by_product:
    qb_by_product[key] = list(dict.fromkeys(qb_by_product[key]))

def get_question_pool_for_product(emp_product):
    if not emp_product:
        return [], "Product not found in Question Bank."
        
    raw_p = emp_product.strip()
    
    # Direct match
    direct_key = get_product_key(raw_p)
    if direct_key in qb_by_product:
        return list(qb_by_product[direct_key]), None

    # Component matching by slash /
    parts = [p.strip() for p in re.split(r'[\/]', raw_p) if p.strip()]
    combined_pool = []
    
    for part in parts:
        pk = get_product_key(part)
        if pk in qb_by_product:
            combined_pool.extend(qb_by_product[pk])
            
    if combined_pool:
        unique_pool = list(dict.fromkeys(combined_pool))
        return unique_pool, None
        
    return [], "Product not found in Question Bank."

# 2. Read Master Employee File
print("Loading Master Employee File...")
master_filepath = "Resource details less tahn 3.5 rating.xlsx"
wb_emp = openpyxl.load_workbook(master_filepath)
sheet_emp = wb_emp.active

emp_rows = list(sheet_emp.iter_rows(values_only=False))
header_cells = emp_rows[0]
header_values = [clean_str(cell.value) for cell in header_cells]

print("Master file header:", header_values)
print("Total rows in Master file (including header):", len(emp_rows))

# Identify Product column index
prod_col_idx = None
for i, h in enumerate(header_values):
    if 'product' in h.lower():
        prod_col_idx = i
        break

if prod_col_idx is None:
    prod_col_idx = 4 # default to index 4 ('Product')

# 3. Create Output Workbook
wb_out = openpyxl.Workbook()
sheet_out = wb_out.active
sheet_out.title = "Employee Question Mapping"

# New Columns to append
new_columns = [
    "Assigned Question 1",
    "Assigned Question 2",
    "Assigned Question 3",
    "Assigned Question 4",
    "Assigned Question 5",
    "Assigned Question 6",
    "Assigned Question 7",
    "Assigned Question 8",
    "Assigned Question 9",
    "Assigned Question 10",
    "Remarks"
]

# Write header row
full_header = header_values + new_columns
sheet_out.append(full_header)

# Process each employee
mapped_count = 0
unmapped_count = 0

for r_idx, row in enumerate(emp_rows[1:], start=2):
    row_values = [cell.value for cell in row]
    emp_prod = clean_str(row_values[prod_col_idx]) if prod_col_idx < len(row_values) else ""
    
    pool, err_remark = get_question_pool_for_product(emp_prod)
    
    if not pool or err_remark:
        # Product not found
        assigned_questions = [""] * 10
        remark = err_remark if err_remark else "Product not found in Question Bank."
        unmapped_count += 1
    else:
        available_count = len(pool)
        # Randomly shuffle question pool
        shuffled_pool = list(pool)
        random.shuffle(shuffled_pool)
        
        if available_count >= 10:
            assigned_questions = shuffled_pool[:10]
            remark = ""
        else:
            assigned_questions = shuffled_pool + [""] * (10 - available_count)
            remark = f"Only {available_count} questions available for this product."
        mapped_count += 1

    new_row = row_values + assigned_questions + [remark]
    sheet_out.append(new_row)

output_filename = "Employee_Question_Mapping.xlsx"
wb_out.save(output_filename)
print(f"Saved mapping output to {output_filename}")
print(f"Successfully mapped {mapped_count} employees.")
print(f"Unmapped (not found) {unmapped_count} employees.")
print(f"Total output rows (excluding header): {sheet_out.max_row - 1}")

# 4. Run Validation Checklist
print("\n=== RUNNING VALIDATION CHECKLIST ===")
val_wb = openpyxl.load_workbook(output_filename, data_only=True)
val_sheet = val_wb.active

val_rows = list(val_sheet.iter_rows(values_only=True))
val_header = val_rows[0]
val_data_rows = val_rows[1:]

assert len(val_data_rows) == len(emp_rows) - 1, f"Row count mismatch! Expected {len(emp_rows) - 1}, got {len(val_data_rows)}"
print(f"PASSED: Row count matches master file exactly ({len(val_data_rows)} rows).")

# Check duplicate questions per employee
dup_q_count = 0
for idx, r in enumerate(val_data_rows):
    qs = [str(q).strip() for q in r[len(header_values):len(header_values)+10] if q and str(q).strip()]
    if len(qs) != len(set(qs)):
        dup_q_count += 1

assert dup_q_count == 0, f"Found {dup_q_count} employees with duplicate assigned questions!"
print("PASSED: Zero duplicate questions for any employee.")

print("All validation checks passed successfully!")
