import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

wb = openpyxl.load_workbook("Question Bank-20th July '26.xlsx", data_only=True)

for sheetname in wb.sheetnames:
    sheet = wb[sheetname]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        continue
    header = [str(c).strip() if c else '' for c in rows[0]]
    q_col = None
    for i, h in enumerate(header):
        if 'question' in h.lower():
            q_col = i
            break
    qs = []
    if q_col is not None:
        for r in rows[1:]:
            if q_col < len(r) and r[q_col]:
                qs.append(str(r[q_col]).strip())
    unique_qs = list(dict.fromkeys(qs))
    print(f"Sheet '{sheetname}': total rows={len(rows)-1}, total q_cell_values={len(qs)}, UNIQUE questions={len(unique_qs)}")
