import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
import re

wb_qb = openpyxl.load_workbook("Question Bank-20th July '26.xlsx", data_only=True)

print("=== QUESTION BANK ANALYSIS ===")
qb_pools = {} # normalized_prod_key -> list of questions

def clean_str(val):
    if val is None:
        return ""
    return str(val).strip()

def normalize_key(s):
    if not s:
        return ""
    # strip, lowercase, replace punctuation/slashes/hyphens with space or standardize
    s = s.strip().lower()
    # replace slashes, hyphens, underscores, spaces with single space
    s = re.sub(r'[\/\-_\s]+', ' ', s)
    return s.strip()

for sheetname in wb_qb.sheetnames:
    sheet = wb_qb[sheetname]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        continue
    header = [clean_str(c) for c in rows[0]]
    
    # find prod col and q col
    prod_col = None
    q_col = None
    for i, h in enumerate(header):
        h_lower = h.lower()
        if 'product' in h_lower and prod_col is None:
            prod_col = i
        if 'question' in h_lower and q_col is None:
            q_col = i
            
    q_count = 0
    prods_seen = set()
    for r in rows[1:]:
        q_val = clean_str(r[q_col]) if q_col is not None and q_col < len(r) else ""
        if not q_val:
            continue
        prod_val = clean_str(r[prod_col]) if prod_col is not None and prod_col < len(r) else ""
        if not prod_val:
            prod_val = sheetname
            
        prods_seen.add(prod_val)
        
        # Add question to pool for this prod_val
        key1 = normalize_key(prod_val)
        if key1 not in qb_pools:
            qb_pools[key1] = []
        qb_pools[key1].append(q_val)
        
        # Also if sheetname differs and offers another product name
        key2 = normalize_key(sheetname)
        if key2 != key1:
            if key2 not in qb_pools:
                qb_pools[key2] = []
            qb_pools[key2].append(q_val)

print(f"Total QB Pools Keys: {len(qb_pools.keys())}")
for k in sorted(qb_pools.keys()):
    print(f"  Pool '{k}': {len(qb_pools[k])} questions")

print("\n=== MASTER EMP PRODUCTS MATCHING ===")
wb_emp = openpyxl.load_workbook("Resource details less tahn 3.5 rating.xlsx", data_only=True)
sheet_emp = wb_emp.active
emp_rows = list(sheet_emp.iter_rows(values_only=True))
header_emp = emp_rows[0]

emp_prods_counts = {}
for r in emp_rows[1:]:
    p = clean_str(r[4]) # Product column index 4
    emp_prods_counts[p] = emp_prods_counts.get(p, 0) + 1

print(f"Total Unique Product Strings in Master File: {len(emp_prods_counts)}")

found_count = 0
not_found_count = 0

for p, count in sorted(emp_prods_counts.items(), key=lambda x: x[0]):
    norm_p = normalize_key(p)
    # Direct match or partial component match (e.g. multi-products like "HLR-HSS/UDM")
    if norm_p in qb_pools:
        print(f"EXACT MATCH: '{p}' (norm: '{norm_p}') -> {len(qb_pools[norm_p])} q's [{count} emps]")
        found_count += count
    else:
        # Check if components separated by / or - match any product
        # e.g., HLR-HSS, EIR/UDM, etc.
        print(f"NO DIRECT MATCH: '{p}' (norm: '{norm_p}') [{count} emps]")
        not_found_count += count

print(f"\nSummary: {found_count} emps matched directly, {not_found_count} emps not matched directly out of {len(emp_rows)-1}")
