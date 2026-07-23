import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

# Load Master File
wb_master = openpyxl.load_workbook("Resource details less tahn 3.5 rating.xlsx", data_only=True)
sheet_master = wb_master.active
master_rows = list(sheet_master.iter_rows(values_only=True))
master_header = master_rows[0]
master_data = master_rows[1:]

# Load Generated Output File
wb_out = openpyxl.load_workbook("Employee_Question_Mapping.xlsx", data_only=True)
sheet_out = wb_out.active
out_rows = list(sheet_out.iter_rows(values_only=True))
out_header = out_rows[0]
out_data = out_rows[1:]

print("=== DOUBLE CHECKING EMPLOYEE INTEGRITY ===")
print(f"Master file row count (excluding header): {len(master_data)}")
print(f"Output file row count (excluding header): {len(out_data)}")

assert len(master_data) == len(out_data), f"Mismatch! Master has {len(master_data)} rows, Output has {len(out_data)} rows."

# Verify row by row matching for Emp ID, EmpName, Role, Domain, Product, Customer, Nokia Email ID, DDH
mismatches = []
master_emp_ids = []
out_emp_ids = []

for idx, (m_row, o_row) in enumerate(zip(master_data, out_data), start=1):
    m_emp_id = m_row[0]
    o_emp_id = o_row[0]
    master_emp_ids.append(m_emp_id)
    out_emp_ids.append(o_emp_id)
    
    # Check all original columns match exactly
    for col_i in range(8):
        m_val = str(m_row[col_i]).strip() if m_row[col_i] is not None else ""
        o_val = str(o_row[col_i]).strip() if o_row[col_i] is not None else ""
        if m_val != o_val:
            mismatches.append((idx, col_i, master_header[col_i], m_val, o_val))

print(f"Total rows checked: {len(master_data)}")
print(f"Total field mismatches found: {len(mismatches)}")

# Check uniqueness and set equality of Emp IDs
set_master_ids = set(master_emp_ids)
set_out_ids = set(out_emp_ids)

print(f"Unique Emp IDs in Master file: {len(set_master_ids)}")
print(f"Unique Emp IDs in Output file: {len(set_out_ids)}")

extra_in_output = set_out_ids - set_master_ids
missing_in_output = set_master_ids - set_out_ids

print(f"Extra Emp IDs in output (not in master): {len(extra_in_output)}")
print(f"Missing Emp IDs in output (in master but missing): {len(missing_in_output)}")

if len(mismatches) == 0 and len(extra_in_output) == 0 and len(missing_in_output) == 0:
    print("\nVERIFICATION PASSED: Questions are assigned ONLY to the exact employees listed in 'Resource details less tahn 3.5 rating.xlsx', with 100% 1-to-1 row alignment!")
else:
    print("\nVERIFICATION FAILED! Details above.")
