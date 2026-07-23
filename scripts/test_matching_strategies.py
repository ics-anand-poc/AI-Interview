import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
import re

def clean_str(val):
    if val is None:
        return ""
    return str(val).strip()

wb_qb = openpyxl.load_workbook("Question Bank-20th July '26.xlsx", data_only=True)

# Build question pools per product
# QB Product names in QB sheets:
# Sheet AAA -> AAA
# Sheet CSD -> CSD
# Sheet NPC -> NPC
# Sheet NCC -> NCC
# Sheet CMG -> CMG
# Sheet CMM -> CMM
# Sheet NRD -> NRD
# Sheet SDL Deployment and upgrade -> SDL (Product Name col says SDL)
# Sheet HLR_HSS CNF Deployment -> HLR/HSS (Product Name col says HLR/HSS)
# Sheet UDM -> AUSF/UDM (Product Name col says AUSF/UDM)
# Sheet CFX -> CFX
# Sheet MRF -> MRF
# Sheet NN -> NN
# Sheet CBIS -> CBIS
# Sheet NCOM -> NCOM
# Sheet NCP -> NCP
# Sheet CBAM -> CBAM
# Sheet NCD -> NCD

qb_questions_by_canon = {} # canonical_name -> list of questions

def to_canon(name):
    if not name:
        return ""
    s = name.upper().strip()
    s = re.sub(r'[^A-Z0-9]+', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    # Normalize common variations
    # HLR HSS vs HSS HLR
    if s in ['HLR HSS', 'HSS HLR', 'HLR_HSS', 'HLR/HSS', 'HLRHSS', 'HSSHLR']:
        return 'HLR-HSS'
    if s in ['AUSF UDM', 'UDM']:
        return 'UDM'
    if s in ['SDL DEPLOYMENT AND UPGRADE']:
        return 'SDL'
    if s in ['ONE NDS', 'NDS']:
        return 'NDS'
    return s

# Read all QB sheets
for sheetname in wb_qb.sheetnames:
    sheet = wb_qb[sheetname]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        continue
    header = [clean_str(c) for c in rows[0]]
    prod_col = None
    q_col = None
    for i, h in enumerate(header):
        h_lower = h.lower()
        if 'product' in h_lower and prod_col is None:
            prod_col = i
        if 'question' in h_lower and q_col is None:
            q_col = i
            
    for r in rows[1:]:
        q_val = clean_str(r[q_col]) if q_col is not None and q_col < len(r) else ""
        if not q_val:
            continue
        prod_val = clean_str(r[prod_col]) if prod_col is not None and prod_col < len(r) else ""
        if not prod_val:
            prod_val = sheetname
            
        canon_p = to_canon(prod_val)
        canon_s = to_canon(sheetname)
        
        # Add to canon_p
        if canon_p not in qb_questions_by_canon:
            qb_questions_by_canon[canon_p] = []
        qb_questions_by_canon[canon_p].append(q_val)
        
        # Also map sheet name canon if different
        if canon_s != canon_p:
            if canon_s not in qb_questions_by_canon:
                qb_questions_by_canon[canon_s] = []
            qb_questions_by_canon[canon_s].append(q_val)

print("Canonical QB Pools:")
for k, v in qb_questions_by_canon.items():
    print(f"  {k}: {len(v)} questions")

# Now check Master Emp Products
wb_emp = openpyxl.load_workbook("Resource details less tahn 3.5 rating.xlsx", data_only=True)
sheet_emp = wb_emp.active
emp_rows = list(sheet_emp.iter_rows(values_only=True))

emp_prods = set()
for r in emp_rows[1:]:
    p = clean_str(r[4])
    if p:
        emp_prods.add(p)

print("\n--- MATCHING RESULTS WITH CANONICAL & COMPONENT MATCHING ---")

matched_count = 0
unmatched_count = 0

for ep in sorted(list(emp_prods)):
    # Try exact canon
    c_ep = to_canon(ep)
    
    # Also try component splitting by / or -
    # e.g., "EIR / HLR-HSS/UDM" -> ["EIR", "HLR-HSS", "UDM"]
    raw_components = re.split(r'[\/]', ep)
    matched_pool = []
    matched_sources = []
    
    if c_ep in qb_questions_by_canon:
        matched_pool = qb_questions_by_canon[c_ep]
        matched_sources.append(c_ep)
    else:
        # try matching individual component products
        for comp in raw_components:
            canon_comp = to_canon(comp)
            if canon_comp in qb_questions_by_canon:
                matched_pool.extend(qb_questions_by_canon[canon_comp])
                matched_sources.append(canon_comp)
                
    # Deduplicate questions in pool if needed
    unique_q_pool = list(dict.fromkeys(matched_pool))
    
    count_for_p = sum(1 for r in emp_rows[1:] if clean_str(r[4]) == ep)
    if unique_q_pool:
        matched_count += count_for_p
        print(f"MATCH: '{ep}' -> sources={matched_sources}, total available q's={len(unique_q_pool)} ({count_for_p} emps)")
    else:
        unmatched_count += count_for_p
        print(f"NO MATCH: '{ep}' ({count_for_p} emps)")

print(f"\nTotal Matched: {matched_count} / {len(emp_rows)-1}")
print(f"Total Unmatched: {unmatched_count} / {len(emp_rows)-1}")
