import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
import re
import random

def clean_str(val):
    if val is None:
        return ""
    return str(val).strip()

def normalize_token(s):
    if not s:
        return ""
    s = s.upper().strip()
    s = re.sub(r'[^A-Z0-9]+', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s

wb_qb = openpyxl.load_workbook("Question Bank-20th July '26.xlsx", data_only=True)

# Question pools by canonical product key
# Recognized product keys:
# AAA, CSD, NPC, NCC, CMG, CMM, NRD, SDL, HLR-HSS, UDM, CFX, MRF, NN, CBIS, NCOM, NCP, CBAM, NCD

qb_by_product = {}

def get_product_key(raw_name, sheet_name=""):
    t_raw = normalize_token(raw_name)
    t_sheet = normalize_token(sheet_name)
    
    # Direct mappings
    for t in [t_raw, t_sheet]:
        if t in ['HLR HSS', 'HSS HLR', 'HLR HSS CNF DEPLOYMENT', 'HLR HSS CNF', 'HSS', 'HLR']:
            return 'HLR-HSS'
        if t in ['AUSF UDM', 'UDM']:
            return 'UDM'
        if t in ['SDL DEPLOYMENT AND UPGRADE', 'SDL']:
            return 'SDL'
        if t in ['AAA', 'CSD', 'NPC', 'NCC', 'CMG', 'CMM', 'NRD', 'CFX', 'MRF', 'NN', 'CBIS', 'NCOM', 'NCP', 'CBAM', 'NCD']:
            return t
    return t_raw if t_raw else t_sheet

for sheetname in wb_qb.sheetnames:
    sheet = wb_qb[sheetname]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        continue
    header = [clean_str(c) for c in rows[0]]
    prod_col = None
    q_col = None
    for i, h in enumerate(header):
        h_lower = h.lower()
        if 'product' in h_lower and prod_col is None:
            prod_col = i
        if 'question' in h_lower and q_col is None:
            q_col = i
            
    for r in rows[1:]:
        q_val = clean_str(r[q_col]) if q_col is not None and q_col < len(r) else ""
        if not q_val:
            continue
        prod_val = clean_str(r[prod_col]) if prod_col is not None and prod_col < len(r) else ""
        
        key = get_product_key(prod_val, sheetname)
        if key not in qb_by_product:
            qb_by_product[key] = []
        qb_by_product[key].append(q_val)

print("Loaded Question Bank pools:")
for k, v in qb_by_product.items():
    print(f"  [{k}]: {len(v)} questions")

# Match function for employee product
def get_questions_for_employee_product(emp_product):
    if not emp_product:
        return [], "Product not found in Question Bank."
        
    raw_p = emp_product.strip()
    norm_p = normalize_token(raw_p)
    
    # 1. Direct canonical match
    direct_key = get_product_key(raw_p)
    if direct_key in qb_by_product:
        pool = list(dict.fromkeys(qb_by_product[direct_key]))
        return pool, None

    # 2. Check components separated by / or -
    # E.g. "EIR/UDM/HLR-HSS" -> ["EIR", "UDM", "HLR-HSS"]
    # E.g. "SDL/UDM" -> ["SDL", "UDM"]
    parts = [p.strip() for p in re.split(r'[\/]', raw_p) if p.strip()]
    combined_pool = []
    matched_components = []
    
    for part in parts:
        pk = get_product_key(part)
        if pk in qb_by_product:
            matched_components.append(pk)
            combined_pool.extend(qb_by_product[pk])
            
    if combined_pool:
        # Deduplicate while preserving order before shuffle
        unique_pool = list(dict.fromkeys(combined_pool))
        return unique_pool, None
        
    return [], "Product not found in Question Bank."

# Read Master File and test
wb_emp = openpyxl.load_workbook("Resource details less tahn 3.5 rating.xlsx", data_only=True)
sheet_emp = wb_emp.active
emp_rows = list(sheet_emp.iter_rows(values_only=True))

matched_emp_count = 0
unmatched_emp_count = 0
emp_prod_summary = {}

for r in emp_rows[1:]:
    emp_id = r[0]
    p_val = clean_str(r[4])
    pool, err = get_questions_for_employee_product(p_val)
    
    if p_val not in emp_prod_summary:
        emp_prod_summary[p_val] = {'count': 0, 'pool_len': len(pool), 'err': err}
    emp_prod_summary[p_val]['count'] += 1
    
    if pool:
        matched_emp_count += 1
    else:
        unmatched_emp_count += 1

print("\n=== SUMMARY OF EMP PRODUCTS ===")
for p, data in sorted(emp_prod_summary.items(), key=lambda x: x[0]):
    status = f"MATCHED ({data['pool_len']} q's available)" if data['pool_len'] > 0 else f"NOT FOUND ({data['err']})"
    print(f"'{p}' ({data['count']} emps) -> {status}")

print(f"\nTOTAL EMPLOYEES: {len(emp_rows)-1}")
print(f"MATCHED EMPLOYEES: {matched_emp_count}")
print(f"UNMATCHED EMPLOYEES: {unmatched_emp_count}")
